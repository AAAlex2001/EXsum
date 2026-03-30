from django.contrib import admin
from unfold.admin import ModelAdmin
from .models import Deal, IncomeExpense, LogOperation, CourseBill, HistoryBill, PercentCourse, ContractorDebtOperation, DealRepayment, CourseHistory
from django.contrib.admin.sites import site
from django.utils.html import escape, mark_safe
from django.conf import settings
from unfold.admin import StackedInline, TabularInline
from .utils import *
from catalog.models import Bill, Currency, Contractor, ContractorHistory
from django.contrib import messages
from .forms import *
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from unfold.decorators import action, display
from django.db.models import QuerySet, Q, OuterRef, Subquery, Sum, Case, When, Value, CharField, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpRequest
from django.shortcuts import redirect
from django.urls import reverse
from catalog.utils import reset_bills_to_initial
from unfold.enums import ActionVariant
from django.contrib.admin import RelatedOnlyFieldListFilter
from django.db.models import Sum
from django.db import transaction
from django import forms
from django.contrib.admin.widgets import RelatedFieldWidgetWrapper
import json
from django.shortcuts import render, redirect
from .forms import *
from django.contrib.admin.utils import model_ngettext
from django.template.response import TemplateResponse
from import_export.admin import ImportExportModelAdmin
from import_export import resources
from import_export.formats import base_formats
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils.html import format_html
from django.utils import timezone
import reversion
from reversion.admin import VersionAdmin
from reversion_compare.admin import CompareVersionAdmin
from decimal import Decimal
from reversion.models import Version
from django.contrib.admin.models import LogEntry, ADDITION
from django.contrib.contenttypes.models import ContentType
from django.forms.models import BaseInlineFormSet
from django.contrib.admin import SimpleListFilter
from django.core.exceptions import ValidationError
from unfold.contrib.filters.admin import RangeDateFilter, RangeDateTimeFilter
from django.db.models.functions import Cast
from django.db.models import Case, When, IntegerField


class NationalCurrencyFilter(admin.SimpleListFilter):
    title = 'Нац.валюта'
    parameter_name = 'national_currency_state'

    def lookups(self, request, model_admin):
        return (
            ('empty', 'Не заполнено'),
            ('filled', 'Заполнено'),
        )

    def queryset(self, request, queryset):

        if self.value() == 'empty':
            return queryset.filter(
                Q(national_currency__isnull=True) | Q(national_currency=0)
            )

        if self.value() == 'filled':
            return queryset.exclude(
                Q(national_currency__isnull=True) | Q(national_currency=0)
            )

        return queryset

class CreatedFilters(SimpleListFilter):
    title = 'Фильтр создателей'
    parameter_name = 'deal_created_users'

    def lookups(self, request, model_admin):
        usernames_list = list(LogEntry.objects.order_by('user__username').values_list('user__username', flat=True).distinct())
        return [(c, c) for c in usernames_list]

        
    def queryset(self, request, queryset):
        if self.value():
            value = self.value()
            ct = ContentType.objects.get_for_model(queryset.model)
            creator_sq = LogEntry.objects.filter(
                content_type=ct,
                object_id=Cast(OuterRef('pk'), output_field=CharField()),
                action_flag=ADDITION
            ).order_by('action_time').values('user__username')[:1]

            queryset = queryset.annotate(
                creator=Subquery(creator_sq)
            ).filter(creator=value)

        return queryset
            

class ManagerFilters(SimpleListFilter):
    title = 'Фильтр кастомные'
    parameter_name = 'deal_filters_v1'

    def lookups(self, request, model_admin):
        RATE_ZERO = 1
        INCOME_NATIONAL_ZERO = 2
        MANAGER_FILTERS = {
            RATE_ZERO: {
                'title': 'Курс прихода 0',
            },
            INCOME_NATIONAL_ZERO: {
                'title': 'Нет суммы прихода и нац.',
            },}
        CHOICE_FILTERS = [(k, v['title']) for k, v in MANAGER_FILTERS.items()]
        filters = set([c for c in CHOICE_FILTERS])
        return [(c) for c in filters]

    def queryset(self, request, queryset):
        if self.value():
            if int(self.value()) == 1:
                in_exes = IncomeExpense.objects.filter(income_account__currency__short_name__in=['USDT', 'USD'], income_rate=0)
                in_exes_list = list(in_exes.values_list('deal_id', flat=True).distinct())
                end_queryset = queryset.filter(id__in=in_exes_list)
                return end_queryset
            if int(self.value()) == 2:
                in_exes = IncomeExpense.objects.exclude(income_account__currency__short_name__in=['USDT', 'USD'])
                in_exes_list = list(in_exes.values_list('deal_id', flat=True).distinct())
                end_queryset = queryset.filter(id__in=in_exes_list, national_currency__isnull=True).filter(Q(cashflow__id__in=[7,9]) | Q(category__id=2))
                return end_queryset
        return queryset





@admin.register(PercentCourse)
class PercentCourseAdmin(admin.ModelAdmin):
    def has_view_permission(self, request, obj=None):
        if request.user.groups.filter(name='restricted').exists():
            return False
        return super().has_view_permission(request, obj=obj)

    def has_change_permission(self, request, obj=None):
        if request.user.groups.filter(name='restricted').exists():
            return False
        return super().has_change_permission(request, obj=obj)


    def has_add_permission(self, request):
        return not PercentCourse.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def get_model_perms(self, request):
        if request.user.is_superuser and request.user.groups.filter(name='restricted').exists():
            return {}
        return super().get_model_perms(request)

@admin.register(LogOperation)
class LogOperationAdmin(ModelAdmin):
	list_display = ['id', 'data',]
	list_display_links = ('id', )

class IncomeExpenseForm(forms.ModelForm):
    class Meta:
        model = IncomeExpense
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for fld in ('income_account', 'expense_account'):
            widget = self.fields[fld].widget
            if isinstance(widget, RelatedFieldWidgetWrapper):
                widget.can_add_related = False
                widget.can_change_related = False
                widget.can_delete_related = False
                # Django >=3.2
                if hasattr(widget, 'can_view_related'):
                    widget.can_view_related = False

reversion.register(IncomeExpense)

class IncomeExpenseInlineFormSet(BaseInlineFormSet):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for form in self.forms:
            ie: IncomeExpense = form.instance
            if not ie.pk:
                form.extra_data = {
                    'income_before': None,
                    'income_after': None,
                    'expense_before': None,
                    'expense_after': None,
                }
                continue
            historys = HistoryBill.objects.filter(in_ex=ie)
            if historys.count() != 0:
                history_obj = historys.last()
                form.extra_data = {
                    'income_before': history_obj.income_before,
                    'income_after': history_obj.income_after,
                    'expense_before': history_obj.expense_before,
                    'expense_after': history_obj.expense_after,
                }
            else:
                form.extra_data = {
                    'income_before': None,
                    'income_after': None,
                    'expense_before': None,
                    'expense_after': None,
                }

class ContractorDebtOperationInline(TabularInline):
    model = ContractorDebtOperation
    form = ContractorDebtOperationForm
    extra = 1
    hide_title = False
    #tab = False

class IncomeExpenseInline(TabularInline):
    model = IncomeExpense
    form = IncomeExpenseForm
    #formset = IncomeExpenseInlineFormSet
    extra = 1
    hide_title = False
    #tab = False



    def delete_model(self, request, obj):
        deal_instance = obj.deal
        super().delete_model(request, obj)
        income_expenses = IncomeExpense.objects.filter(deal=deal_instance)
        for in_ex in income_expenses:
            if not in_ex.rate_calculated:
                income_rate = (in_ex.expense_amount * in_ex.expense_account.rate) / in_ex.income_amount
                rate_calculated = True
                IncomeExpense.objects.filter(id=in_ex.id).update(income_rate=income_rate,
                                                                 rate_calculated=rate_calculated)
        ids = set(list(income_expenses.values_list('income_account__id', flat=True)) + list(
            income_expenses.values_list('expense_account__id', flat=True)))
        accounts = Bill.objects.filter(id__in=ids)
        for bill in accounts:
            total_sum = get_total_rate_balance(bill, 'income_expense_v3')
            messages.success(request, "Sum {0}".format(str(total_sum)))
            bill.remainder = total_sum
            bill.save()

class MyPartResource(resources.ModelResource):
    class Meta:
        model = Deal
        fields = ('id', 'formatted_date_create', 'income_account_1', 'income_amount_1', 'expense_account_1',  'expense_amount_1', 'revenue', 'percent', 'rate_old', 'rate_new', 'contractor', 'comment', 'category', 'cashflow',)

class DealAdminForm(forms.ModelForm):
    class Meta:
        model = Deal
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        closed = cleaned_data.get('closed')
        deal_1c = cleaned_data.get('deal_1c')

        if closed and not deal_1c:
            raise ValidationError({'closed': 'Нельзя закрыть сделку без заполнения поля "Заказ 1с".'})

        return cleaned_data
    
reversion.register(Deal, follow=['deal_data']) 
@admin.register(Deal)
class DealAdmin(CompareVersionAdmin, ModelAdmin):
    #list_filter_sheet = False  # отключаем новую панель фильтров
    list_filter_submit = True
    list_per_page = 50
    follow = ['deal_data']
    #form = DealAdminForm
    list_display = ['id', 'formatted_date_create', 'income_account_1', 'income_amount_1', 'expense_account_1',  'expense_amount_1', 'revenue', 'percent', 'debts', 'debts_name', 'rate_old', 'rate_new', 'contractor', 'comment', 'category', 'cashflow', 'national_currency_custom', 'created_by']
    list_display_links = ('id', 'income_account_1', 'income_amount_1', 'expense_account_1', 'expense_amount_1', 'revenue', 'percent', 'rate_old', 'rate_new', 'contractor', 'comment',  'category', 'cashflow', )


    list_filter = ('category',  'cashflow', 'contractor', 'closed', ('deal_data__income_account', RelatedOnlyFieldListFilter), ('deal_data__expense_account', RelatedOnlyFieldListFilter), ManagerFilters, ("date_create", RangeDateTimeFilter), CreatedFilters, NationalCurrencyFilter)
    ordering = ['-date_create']
    search_fields = ['id', 'comment']
    change_list_template = "bills.html"
    change_form_template = "bills_one.html"

    fieldsets = (
        ('Основные данные', {
            'fields': ("date_create", "category","rate_contractors", "contractor", "cashflow", "closed", "rate", "comment", "duty_deal",
                       "national_currency", "deal_1c", "by_nikita"),
        }),


    )
    inlines = [IncomeExpenseInline, ContractorDebtOperationInline,]

    # actions_list disabled — URL-ы регистрируются через get_urls(), кнопки в bills.html
    # actions_list = ["calculate_rate", "calculate_rate_500"]

    actions_detail = ['history_detail', ]
    actions = ["make_is_closed", "clone_deal", "bulk_update_category_contractor", "export_xlsx"]

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'contractor' in form.base_fields:
            form.base_fields['contractor'].required = False
        return form

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path(
                'calculate-rate/',
                self.admin_site.admin_view(self.calculate_rate),
                name='operation_deal_calculate_rate',
            ),
            path(
                'calculate-rate-500/',
                self.admin_site.admin_view(self.calculate_rate_500),
                name='operation_deal_calculate_rate_500',
            ),
        ]
        return custom_urls + urls



    """def get_queryset(self, request):
        return super().get_queryset(request)"""

    #@action(label='Скачать XLSX')
    def export_xlsx(self, request, queryset):
        import re
        wb = Workbook()
        ws = wb.active
        ws.title = "Deals"
        headers = []
        for field_name in self.list_display:
            admin_fn = getattr(self, field_name, None)
            if callable(admin_fn) and hasattr(admin_fn, 'short_description'):
                headers.append(admin_fn.short_description)
                continue
            model_attr = getattr(self.model, field_name, None)
            if callable(model_attr) and hasattr(model_attr, 'short_description'):
                headers.append(model_attr.short_description)
                continue
            try:
                field = self.model._meta.get_field(field_name)
            except Exception:
                headers.append(field_name.replace('_', ' ').title())
            else:
                headers.append(field.verbose_name.title())

        ws.append(headers)
        for obj in queryset:
            row = []
            for field_name in self.list_display:
                admin_fn = getattr(self, field_name, None)
                if callable(admin_fn) and hasattr(admin_fn, 'short_description'):
                    value = admin_fn(obj)
                else:
                    attr = getattr(obj, field_name, '')
                    if callable(attr):
                        try:
                            value = attr()
                        except TypeError:
                            value = str(attr)
                    else:
                        value = attr
                if not isinstance(value, (int, float, str, bool, type(None))):
                    value = str(value)
                if isinstance(value, str):
                    value = re.sub(r'<.*?>', '', value) 
                row.append(value)

            ws.append(row)
        for i, _ in enumerate(headers, 1):
            col = get_column_letter(i)
            ws.column_dimensions[col].auto_size = True
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=deals.xlsx'
        wb.save(response)
        return response


    export_xlsx.short_description = "Скачать выбранные в XLSX"



    def bulk_update_category_contractor(self, request, queryset):
        if 'apply' in request.POST:
            form = BulkUpdateForm(request.POST)
            if form.is_valid():
                changed = form.changed_data
                category = form.cleaned_data['category']
                contractor = form.cleaned_data['contractor']
                cashflow = form.cleaned_data['cashflow']
                updated = 0
                for obj in queryset:
                    if "category" in changed:
                    #if category is not None:
                        obj.category = category
                    #if contractor is not None:
                    if "contractor" in changed:
                        obj.contractor = contractor
                    #if cashflow is not None:
                    if "cashflow" in changed:
                        obj.cashflow = cashflow
                    obj.save()
                    updated += 1

                self.message_user(request, f"Обновлено {updated} {model_ngettext(self.opts, updated)}.",
                                  messages.SUCCESS)
                return redirect(request.get_full_path())
        else:
            form = BulkUpdateForm()
        context = self.admin_site.each_context(request)
        context.update({
            'form': form,
            'queryset': queryset,
            'title': 'Обновить категорию или контрагента или ДДС',
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
            'opts': self.model._meta,
            'app_label': self.model._meta.app_label,
            'model_name': self.model._meta.model_name,
            'cl': self.get_changelist_instance(request),
            'media': self.media,
            'request': request,
            'action': 'bulk_update_category_contractor',
        })
        return TemplateResponse(
            request,
            "bulk_update_category_contractor.html",
            context,

        )

    bulk_update_category_contractor.short_description = "Изменить категорию / контрагента / ДДС"

    class Media:
        css = {
            'all': (
                'operation/custom_admin.css',
            )
        }
        js = ('operation/custom_admin.js',)

    def get_search_results(self, request, queryset, search_term):
        try:
            search_term_str = str(search_term)
            if ',' in search_term:
                search_term = search_term.replace(',', '.')
            amount = Decimal(search_term)
        except (InvalidOperation, ValueError):
            return super().get_search_results(request, queryset, search_term)
        qs_filtered = queryset.filter(
            Q(deal_data__income_amount=amount) |
            Q(deal_data__expense_amount=amount) |
            Q(comment__icontains=search_term_str) |
            Q(id=int(float(amount))) |
            Q(national_currency=Decimal(float(amount)))
        ).distinct()

        return qs_filtered, True

    def get_list_display(self, request):
        if request.user.is_superuser and request.user.groups.filter(name='restricted').exists():
            if request.user.username == 'Nikita':
                return ('id', 'formatted_date_create', 'income_account_1', 'income_amount_1', 'expense_account_1', 'expense_amount_1', 'category', 'cashflow', 'contractor', 'comment', 'national_currency_custom', 'created_by')
            else:
                return ('id', 'formatted_date_create', 'income_account_1', 'income_amount_1', 'expense_account_1', 'expense_amount_1', 'category', 'cashflow', 'contractor', 'comment', 'national_currency_custom')
        return super().get_list_display(request)

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        if request.user.is_superuser and request.user.groups.filter(name='restricted').exists() and request.user.id != 3:
            extra_context['autocomplete'] = True
        else:
            extra_context['autocomplete'] = False
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        latest_hb = HistoryBill.objects.filter(
            in_ex__deal=OuterRef('pk')
        ).order_by(
            '-in_ex__date_create',
        )
        income_expense_subquery = IncomeExpense.objects.filter(
            deal=OuterRef('pk'),
            income_account__currency__short_name='RUB'  # Фильтрация по валюте
        ).values('income_amount')[:1]
        return qs.annotate(
            my_color_data=Case(
                When(closed=False, then=Value('#FFEBEB')),
                When(category_id=1, then=Value('#F6FFEB')),
                When(category_id=3, then=Value('#FFFBEB')),
                default=Value('#FFFFFF'),
                output_field=CharField(),
            ),
            hb_income_before=Subquery(
                latest_hb.values('income_before')[:1],
                output_field=DecimalField(max_digits=18, decimal_places=8)
            ),
            hb_income_after=Subquery(
                latest_hb.values('income_after')[:1],
                output_field=DecimalField(max_digits=18, decimal_places=8)
            ),
            hb_expense_before=Subquery(
                latest_hb.values('expense_before')[:1],
                output_field=DecimalField(max_digits=18, decimal_places=8)
            ),
            hb_expense_after=Subquery(
                latest_hb.values('expense_after')[:1],
                output_field=DecimalField(max_digits=18, decimal_places=8)
            ),
            national_currency_value=Case(
            # Если значение national_currency не равно 0, то используем его
            When(national_currency__gt=0, then=F('national_currency')),
            # Если national_currency == 0, проверяем валюту счета
            When(
                national_currency=0,
                then=Coalesce(
                    Subquery(income_expense_subquery),
                    Value(0)
                )
            ),
            default=Value(0),
            output_field=DecimalField(max_digits=18, decimal_places=8)
        )
            
        )

    def get_color(self, obj):
        if obj.category:
            if not obj.closed:
                return '#FFEBEB'
            if obj.category.id == 1:
                return '#F6FFEB'
            elif obj.category.id == 4:
                return '#FFFFFF'
            elif obj.category.id == 3:
                return '#FFFBEB'
            elif obj.category.id == 2:
                return '#FFFFFF'
        if not obj.closed:
            return '#FFEBEB'
        return '#FFFFFF'

    def make_is_closed(modeladmin, request, queryset):
        queryset.update(closed=True)

    make_is_closed.short_description = 'Закрыть'

    def clone_deal(modeladmin, request, queryset):
        for deal in queryset:
            original = deal
            data = {
                f.name: getattr(original, f.name)
                for f in Deal._meta.fields
                if f.editable
                   and not f.auto_created
                   and f.name not in ('id', 'carvet_id', 'date_create')
            }
            data['date_create'] = timezone.now()
            with transaction.atomic():
                new_deal = Deal.objects.create(**data)
                for ie in original.deal_data.all():
                    ie_data = {
                        f.name: getattr(ie, f.name)
                        for f in IncomeExpense._meta.fields
                        if f.editable
                           and not f.auto_created
                           and f.name != 'id'
                    }
                    ie_data['deal'] = new_deal
                    IncomeExpense.objects.create(**ie_data)
        return True

    clone_deal.short_description = 'Дублировать'

    @action(description="Пересчитать 500", icon="key", url_path="calculate-rate-500")
    def calculate_rate_500(self, request: HttpRequest):
        from .tasks import recalculate_courses_dutys_500
        recalculate_courses_dutys_500.delay()
        messages.success(request, "Курсы поставлены в очередь(последние 500)")
        return redirect(reverse("admin:operation_deal_changelist"))

    @action(description="Пересчитать", icon="key", url_path="calculate-rate")
    def calculate_rate(self, request: HttpRequest):# request: HttpRequest,queryset: QuerySet,

        from .tasks import recalculate_courses_dutys
        recalculate_courses_dutys.delay()
        messages.success(request, "Курсы и долги поставлены в очередь на пересчет")
        return redirect(reverse("admin:operation_deal_changelist"))

    @action(
        description="История изменений",
        url_path="history-detail",
        attrs={"target": "_blank"},
        permissions=["changeform_action"]
    )
    #def history_detail(self, request: HttpRequest, object_id: int):
    def history_detail(self, request, *args, **kwargs):
        object_id = kwargs.get('object_id') or (args[0] if args else None)
        versions = Version.objects.get_for_object(self.get_object(request, object_id)) \
            .order_by('revision__date_created')
        total = versions.count()
        if len(versions) < 2:
            return redirect(reverse('admin:operation_deal_history', args=[object_id]))
        v1, v2 = versions[total-2], versions[total-1]

        compare_url = (
                reverse('admin:operation_deal_compare', args=[object_id])
                + f'?version_id1={v1.id}&version_id2={v2.id}'
        )
        return redirect(compare_url)




    def get_fieldsets(self, request, obj=None):
        fieldsets = list(super().get_fieldsets(request, obj))

        if request.user.id != 1:
            # Убираем duty_deal для этой группы
            fieldsets[0][1]['fields'] = tuple(
                f for f in fieldsets[0][1]['fields']
                if f != 'duty_deal'
            )

        return fieldsets


    def has_changeform_action_permission(self, request: HttpRequest, object_id: int):
        if request.user.groups.filter(name='restricted').exists():
            return False
        return self.has_change_permission(request, obj=self.get_object(request, object_id))



    #@display(description="Курс прихода", label="success ")
    def income_account_sum(self, obj):
        in_exes =  IncomeExpense.objects.filter(deal=obj, income_account__isnull=False)
        if in_exes.count() != 0:
            in_ex = in_exes.first()
            amount = in_ex.income_amount
            amount_in = '{0} '.format(str(amount.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)))
            return  '{0} - {1}'.format(str(in_ex.income_account), str(amount_in))
        return '-'


    def expense_account_sum(self, obj):
        in_exes = IncomeExpense.objects.filter(deal=obj, expense_account__isnull=False)
        if in_exes.count() != 0:
            in_ex = in_exes.first()
            amount = in_ex.expense_amount
            amount_ex = '{0} '.format(str(amount.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)))
            return '{0} - {1}'.format(str(in_ex.expense_account), str(amount_ex))
        return '-'

    def delete_queryset(self, request, queryset):
        for obj in queryset:
            deal_instance = obj
            income_expenses = IncomeExpense.objects.filter(deal=deal_instance)
            ids = set(list(income_expenses.values_list('income_account__id', flat=True)) + list(
                income_expenses.values_list('expense_account__id', flat=True)))
            super().delete_model(request, obj)
            accounts = Bill.objects.filter(id__in=ids)
            messages.success(request, "Bils count {0}".format(str(accounts.count())))
            for bill in accounts:
                total_sum = get_total_rate_balance(bill, 'income_expense_v3')
                messages.success(request, "Sum {0}".format(str(total_sum)))
                bill.remainder = total_sum
                bill.save()

    def delete_model(self, request, obj):
        deal_instance = obj
        income_expenses = IncomeExpense.objects.filter(deal=deal_instance)
        ids = set(list(income_expenses.values_list('income_account__id', flat=True)) + list(
            income_expenses.values_list('expense_account__id', flat=True)))
        super().delete_model(request, obj)
        accounts = Bill.objects.filter(id__in=ids)
        messages.success(request, "Bilss count {0}".format(str(accounts.count())))
        for bill in accounts:
            total_sum = get_total_rate_balance(bill, 'income_expense_v3')
            messages.success(request, "Sum {0}".format(str(total_sum)))
            bill.remainder = total_sum
            bill.save()


    def save_related(self, request, form, formsets, change):
        from operation.utils import apply_calcs_to_new_duty, apply_calc_to_existing_duties,detect_status_ie
        if not change:
            super().save_related(request, form, formsets, change)
            deal_instance = form.instance
            income_expenses = IncomeExpense.objects.filter(deal=deal_instance)
            duty = 0
            for in_ex in income_expenses:
                rate_calculated = False
                if not in_ex.rate_calculated:
                    if in_ex.expense_amount and in_ex.expense_account and in_ex.income_amount:
                        income_rate = (in_ex.expense_amount * in_ex.expense_account.rate) / in_ex.income_amount
                        if not in_ex.income_rate:
                            IncomeExpense.objects.filter(id=in_ex.id).update(income_rate=income_rate)
                            rate_calculated = True
                    if in_ex.expense_account:
                        expense_rate = in_ex.expense_account.rate
                        if not in_ex.expense_rate:
                            IncomeExpense.objects.filter(id=in_ex.id).update(expense_rate=expense_rate)
                    IncomeExpense.objects.filter(id=in_ex.id).update(rate_calculated=rate_calculated)
                if deal_instance.contractor:
                    if in_ex.income_account and not in_ex.expense_account:
                        if in_ex.income_account.currency.short_name in ['USD', 'USDT']:
                            duty += in_ex.income_amount * in_ex.income_rate
                        else:
                            duty += in_ex.income_amount
                    if in_ex.expense_account and not in_ex.income_account:
                        if in_ex.expense_account.currency.short_name in ['USD', 'USDT']:
                            duty -= in_ex.expense_amount * deal_instance.rate
                        else:
                            duty -= in_ex.expense_amount
            debts = ContractorDebtOperation.objects.filter(deal=deal_instance)
            nation_debs = debts.filter(operation_type='write_off', currency__short_name='RUB')
            usdt_debs = debts.filter(operation_type='write_on', currency__short_name='USDT')
            if nation_debs.count() != 0 and usdt_debs.count() != 0:
                Deal.objects.filter(id=deal_instance.id).update(rate_contractors=Currency.objects.get(short_name='USDT').rate)
            if deal_instance.duty_deal != 0:
                duty -= deal_instance.duty_deal
            if duty != 0 and deal_instance.contractor and deal_instance.closed:
                history_obj_olds = ContractorHistory.objects.filter(deal=deal_instance)
                if history_obj_olds.count() != 0:
                    history_obj_olds.delete()
                if ContractorHistory.objects.filter(contractor=deal_instance.contractor).count() != 0:
                    last_hist = ContractorHistory.objects.filter(contractor=deal_instance.contractor).last()
                    duty += last_hist.duty
                    history_obj = ContractorHistory(contractor=deal_instance.contractor, deal=deal_instance)
                    history_obj.duty = Decimal(duty)
                    history_obj.duty_usdt = last_hist.duty_usdt
                    history_obj.save()
                else:
                    #duty += last_hist.duty
                    history_obj = ContractorHistory(contractor=deal_instance.contractor, deal=deal_instance)
                    duty += deal_instance.contractor.duty
                    history_obj.duty = Decimal(duty)
                    history_obj.duty_usdt = deal_instance.contractor.duty_usdt
                    history_obj.save()
            ids = set(list(income_expenses.values_list('income_account__id', flat=True)) + list(income_expenses.values_list('expense_account__id', flat=True)))
            accounts = Bill.objects.filter(id__in=ids)
            for bill in accounts:
                total_sum = get_total_rate_balance(bill, 'income_expense_v3')
                messages.success(request, "Sum {0}".format(str(total_sum)))
                #bill.rate = total_rate
                bill.remainder = total_sum
                bill.save()
            if income_expenses.filter(income_account__currency__short_name__in=['USD', 'USDT']).count() != 0 and deal_instance.closed:
                if income_expenses.filter(income_account__currency__short_name='USD').count() != 0:
                    old_rate, currency_rate = update_currency_rate_v3('USD')
                    c_b = CourseBill(deal=deal_instance)
                    if IncomeExpense.objects.filter(income_account__currency__short_name='USD').count() > 1:
                        c_b.rate_old = old_rate
                    c_b.rate_new = currency_rate
                    c_b.save()
                    messages.success(request, "Currency {0} rate {1}".format("USD", str(currency_rate)))
                elif income_expenses.filter(income_account__currency__short_name='USDT').count() != 0:
                    old_rate, currency_rate = update_currency_rate_v3('USDT')
                    c_b = CourseBill(deal=deal_instance)
                    if IncomeExpense.objects.filter(income_account__currency__short_name='USDT').count() > 1:
                        c_b.rate_old = old_rate
                    c_b.rate_new = currency_rate
                    c_b.save()
                    messages.success(request, "Currency {0} rate {1}".format("USDT", str(currency_rate)))
            recalculate_remainders_v3()
            # contragent
            if deal_instance.closed:
                debt_objs = ContractorDebtOperation.objects.filter(deal=deal_instance)
                for debt in debt_objs:
                    contractor = debt.contractor
                    amount = debt.amount
                    currency_name = debt.currency.short_name
                    percent = debt.percent
                    operation_type = debt.operation_type
                    duty = 0
                    duty_usdt = 0
                    if currency_name == 'USDT':
                        if ContractorHistory.objects.filter(contractor=contractor).count() != 0:
                            last_hist = ContractorHistory.objects.filter(contractor=contractor).last()
                            duty_usdt += last_hist.duty_usdt
                            duty += last_hist.duty
                            if operation_type == 'write_off':
                                duty_usdt -= amount
                            else:
                                duty_usdt += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
                        else:
                            duty += contractor.duty
                            duty_usdt += contractor.duty_usdt
                            if operation_type == 'write_off':
                                duty_usdt -= amount
                            else:
                                duty_usdt += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
                    else:
                        if ContractorHistory.objects.filter(contractor=contractor).count() != 0:
                            last_hist = ContractorHistory.objects.filter(contractor=contractor).last()
                            duty_usdt += last_hist.duty_usdt
                            duty += last_hist.duty
                            if operation_type == 'write_off':
                                duty -= amount
                            else:
                                duty += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
                        else:
                            duty += contractor.duty
                            duty_usdt += contractor.duty_usdt
                            if operation_type == 'write_off':
                                duty -= amount
                            else:
                                duty += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
            # end contragent
            # FIFO
            deal = deal_instance
            if deal.contractor and deal.closed and (deal.contractor.id == 1 or ContractorDebtOperation.objects.filter(contractor__id=1, operation_type='write_on').count() != 0):
                contractor = Contractor.objects.get(id=1)
                contractor_duty_total = abs(contractor.duty)
                contractor_duty_repaid = 0
                try:
                    in_exes = IncomeExpense.objects.filter(deal=deal)
                    debts = ContractorDebtOperation.objects.filter(deal=deal, contractor=contractor)
                    remainder = 0
                    if in_exes.count() != 0:
                        for in_ex in in_exes:
                            if in_ex.income_account:
                                if in_ex.income_account.currency.short_name in ['USD', 'USDT']:
                                    remainder += in_ex.income_amount * in_ex.income_rate
                                else:
                                    remainder += in_ex.income_amount
                    elif debts.count() != 0:
                        for debt in debts:
                            remainder += debt.amount
                except:
                    remainder = 0
                if contractor.duty < 0 and contractor_duty_repaid < contractor_duty_total and remainder > 0:
                    need = contractor_duty_total - contractor_duty_repaid
                    pay = min(need, remainder)
                    if pay > 0:
                        DealRepayment.objects.create(
                            calc=deal,
                            contractor_duty=True,
                            amount=pay,
                            date_create=deal.date_create
                        )
                    contractor_duty_repaid += pay
                    remainder -= pay
                    deal.status_ie = 2 if contractor_duty_repaid == contractor_duty_total else 3
                    deal.save()
                expense_deals = list(
                    Deal.objects.filter(contractor=contractor, closed=True)
                    .filter(deal_data__income_account__isnull=True)
                    .order_by('date_create')
                )

                # ===== Добавляем сделки без IncomeExpense, но с duty_deal =====
                duty_only_deals = list(
                    Deal.objects.filter(contractor=contractor, closed=True)
                    .filter(deal_data__isnull=True, duty_deal__isnull=False)
                    .order_by('date_create')
                )

                # Объединяем все расходные сделки и сортируем по дате

                all_expense_deals = list({d.id: d for d in list(expense_deals) + list(duty_only_deals)}.values())
                all_expense_deals.sort(key=lambda d: d.date_create)
                # ===== 2. Гасим все расходные сделки FIFO =====
                for expense_deal in all_expense_deals:
                    if remainder <= 0:
                        break

                    # Проверяем, есть ли уже погашение для этого дохода и расхода
                    if DealRepayment.objects.filter(calc=deal, duty=expense_deal).exists():
                        continue

                    # Получаем сумму расходной сделки
                    in_ex_exp = expense_deal.deal_data.first() if expense_deal.deal_data.exists() else None
                    if in_ex_exp:
                        if in_ex_exp.expense_account.currency.short_name in ['USD', 'USDT']:
                            expense_amount = in_ex_exp.expense_amount * expense_deal.rate
                        else:
                            expense_amount = in_ex_exp.expense_amount
                    elif getattr(expense_deal, 'duty_deal', None):
                        expense_amount = expense_deal.duty_deal
                    else:
                        continue

                    already_repaid = expense_deal.amount_repaid or 0
                    need = expense_amount - already_repaid
                    if need <= 0:
                        continue

                    pay = min(need, remainder)
                    if pay > 0:
                        DealRepayment.objects.create(
                            calc=deal,
                            duty=expense_deal,
                            amount=pay,
                            date_create=deal.date_create
                        )

                    # Обновляем статус расходной сделки
                    expense_deal.amount_repaid = already_repaid + pay
                    if expense_deal.amount_repaid >= expense_amount:
                        expense_deal.is_repaid = True
                        expense_deal.status_ie = 2
                    else:
                        expense_deal.status_ie = 1
                    expense_deal.save()

                    remainder -= pay

                # Если приход полностью распределен
                if remainder == 0:
                    deal.status_ie = 3
                    deal.save()

            # End fifo

        else:
            super().save_related(request, form, formsets, change)
            import time
            deal_instance = form.instance
            if deal_instance.closed and "closed" in form.changed_data:
                from .tasks import recalculate_courses_dutys
                messages.success(request, 'Пересчет отправлен в фоновоую задачу')
                time.sleep(4)
                #recalculate_courses_dutys.delay()
                #self.calculate_rate(request)
                # contragent
                debt_objs = ContractorDebtOperation.objects.filter(deal=deal_instance)
                for debt in debt_objs:
                    contractor = debt.contractor
                    amount = debt.amount
                    currency_name = debt.currency.short_name
                    percent = debt.percent
                    operation_type = debt.operation_type
                    duty = 0
                    duty_usdt = 0
                    if currency_name == 'USDT':
                        if ContractorHistory.objects.filter(contractor=contractor).count() != 0:
                            last_hist = ContractorHistory.objects.filter(contractor=contractor).last()
                            duty_usdt += last_hist.duty_usdt
                            duty += last_hist.duty
                            if operation_type == 'write_off':
                                duty_usdt -= amount
                            else:
                                duty_usdt += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
                        else:
                            duty += contractor.duty
                            duty_usdt += contractor.duty_usdt
                            if operation_type == 'write_off':
                                duty_usdt -= amount
                            else:
                                duty_usdt += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
                    else:
                        if ContractorHistory.objects.filter(contractor=contractor).count() != 0:
                            last_hist = ContractorHistory.objects.filter(contractor=contractor).last()
                            duty_usdt += last_hist.duty_usdt
                            duty += last_hist.duty
                            if operation_type == 'write_off':
                                duty -= amount
                            else:
                                duty += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
                        else:
                            duty += contractor.duty
                            duty_usdt += contractor.duty_usdt
                            if operation_type == 'write_off':
                                duty -= amount
                            else:
                                duty += amount
                            history_obj = ContractorHistory(contractor=contractor, deal=deal_instance)
                            history_obj.duty = Decimal(duty)
                            history_obj.duty_usdt = Decimal(duty_usdt)
                            history_obj.save()
                # end contragent
                # FIFO
                deal = deal_instance
                contractor = Contractor.objects.get(id=1)
                if deal.contractor and deal.closed and (deal.contractor.id == 1 or ContractorDebtOperation.objects.filter(contractor__id=1, operation_type='write_on').count() != 0):
                    contractor_duty_total = abs(contractor.duty)
                    contractor_duty_repaid = 0
                    try:
                        in_exes = IncomeExpense.objects.filter(deal=deal)
                        debts = ContractorDebtOperation.objects.filter(deal=deal, contractor=contractor)
                        remainder = 0
                        if in_exes.count() != 0:
                            for in_ex in in_exes:
                                if in_ex.income_account:
                                    if in_ex.income_account.currency.short_name in ['USD', 'USDT']:
                                        remainder += in_ex.income_amount * in_ex.income_rate
                                    else:
                                        remainder += in_ex.income_amount
                        elif debts.count() != 0:
                            for debt in debts:
                                remainder += debt.amount
                    except:
                        remainder = 0
                    if contractor.duty < 0 and contractor_duty_repaid < contractor_duty_total and remainder > 0:
                        need = contractor_duty_total - contractor_duty_repaid
                        pay = min(need, remainder)
                        if pay > 0:
                            DealRepayment.objects.create(
                                calc=deal,
                                contractor_duty=True,
                                amount=pay,
                                date_create=deal.date_create
                            )
                        contractor_duty_repaid += pay
                        remainder -= pay
                        deal.status_ie = 2 if contractor_duty_repaid == contractor_duty_total else 3
                        deal.save()
                    expense_deals = list(
                        Deal.objects.filter(contractor=contractor, closed=True)
                        .filter(deal_data__income_account__isnull=True)
                        .order_by('date_create')
                    )

                    # ===== Добавляем сделки без IncomeExpense, но с duty_deal =====
                    duty_only_deals = list(
                        Deal.objects.filter(contractor=contractor, closed=True)
                        .filter(deal_data__isnull=True, duty_deal__isnull=False)
                        .order_by('date_create')
                    )

                    # Объединяем все расходные сделки и сортируем по дате

                    all_expense_deals = list({d.id: d for d in list(expense_deals) + list(duty_only_deals)}.values())
                    all_expense_deals.sort(key=lambda d: d.date_create)
                    # ===== 2. Гасим все расходные сделки FIFO =====
                    for expense_deal in all_expense_deals:
                        if remainder <= 0:
                            break

                        # Проверяем, есть ли уже погашение для этого дохода и расхода
                        if DealRepayment.objects.filter(calc=deal, duty=expense_deal).exists():
                            continue

                        # Получаем сумму расходной сделки
                        in_ex_exp = expense_deal.deal_data.first() if expense_deal.deal_data.exists() else None
                        if in_ex_exp:
                            if in_ex_exp.expense_account.currency.short_name in ['USD', 'USDT']:
                                expense_amount = in_ex_exp.expense_amount * expense_deal.rate
                            else:
                                expense_amount = in_ex_exp.expense_amount
                        elif getattr(expense_deal, 'duty_deal', None):
                            expense_amount = expense_deal.duty_deal
                        else:
                            continue

                        already_repaid = expense_deal.amount_repaid or 0
                        need = expense_amount - already_repaid
                        if need <= 0:
                            continue

                        pay = min(need, remainder)
                        if pay > 0:
                            DealRepayment.objects.create(
                                calc=deal,
                                duty=expense_deal,
                                amount=pay,
                                date_create=deal.date_create
                            )

                        # Обновляем статус расходной сделки
                        expense_deal.amount_repaid = already_repaid + pay
                        if expense_deal.amount_repaid >= expense_amount:
                            expense_deal.is_repaid = True
                            expense_deal.status_ie = 2
                        else:
                            expense_deal.status_ie = 1
                        expense_deal.save()

                        remainder -= pay

                    # Если приход полностью распределен
                    if remainder == 0:
                        deal.status_ie = 3
                        deal.save()

            # End fifo



    def get_bills(self):
        bills = (
            Bill.objects
            .filter(status=True)
            .order_by(
                Case(
                    When(id=8, then=0),
                    When(id=22, then=1),
                    default=2,
                    output_field=IntegerField(),
                ),
                'id'
            )
        )
        return  bills

    def get_bills_amount(self):
        bills = Bill.objects.all()
        total = 0
        for bill in bills:
            if bill.currency.short_name in ['USD', 'USDT']:
                total += bill.rate * bill.remainder
            else:
                total += bill.remainder
        in_exes = IncomeExpense.objects.filter(deal__closed=False)
        sum_unclosed = 0
        for in_ex in in_exes:
            if in_ex.income_account:
                if in_ex.income_account.currency.short_name in ['USD', 'USDT']:
                    sum_unclosed -= in_ex.income_amount * in_ex.income_rate
                else:
                    sum_unclosed -= in_ex.income_amount
            if in_ex.expense_account:
                if in_ex.expense_account.currency.short_name in ['USD', 'USDT']:
                    sum_unclosed += in_ex.expense_amount * in_ex.expense_rate
                else:
                    sum_unclosed += in_ex.income_amount
        total = total - sum_unclosed
        return Decimal(total).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

    def get_bills_amount_admin(self):
        try:
            import requests
            url = 'https://api.rapira.net/open/market/rates'
            r = requests.get(url, timeout=2)
            r.raise_for_status()
            data = r.json()
            course_percent = PercentCourse.objects.first().value
            course_main = Decimal(data['data'][0]['close'])
            course_5 = course_main * (Decimal(1 + (course_percent / 100)))
            bills = Bill.objects.all()
            total = 0
            for bill in bills:
                if bill.currency.short_name in ['USD', 'USDT']:
                    total += bill.remainder * course_5
                else:
                    total += bill.remainder
            in_exes = IncomeExpense.objects.filter(deal__closed=False)
            sum_unclosed = 0
            for in_ex in in_exes:
                if in_ex.income_account:
                    if in_ex.income_account.currency.short_name in ['USD', 'USDT']:
                        sum_unclosed -= in_ex.income_amount * in_ex.income_rate
                    else:
                        sum_unclosed -= in_ex.income_amount
                if in_ex.expense_account:
                    if in_ex.expense_account.currency.short_name in ['USD', 'USDT']:
                        sum_unclosed += in_ex.expense_amount * in_ex.expense_rate
                    else:
                        sum_unclosed += in_ex.income_amount
            total = total - sum_unclosed
            return total.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        except:
            return 0

    def get_sum_obligation(self):
        latest_history = ContractorHistory.objects.filter(
            contractor=OuterRef('pk')
        ).order_by('-date_create')

        contractors = Contractor.objects.annotate(
            last_duty_usdt=Subquery(
                latest_history.values('duty_usdt')[:1],
                output_field=DecimalField()
            )
        )

        total = contractors.aggregate(
            total=Coalesce(
                Sum('last_duty_usdt'),
                Decimal('0'),
                output_field=DecimalField()
            )
        )['total']

        return total.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

    def get_sum_usdt(self):
        total = Bill.objects.filter(
            currency__short_name='USDT'
        ).aggregate(
            total_remainder=Sum('remainder')
        )['total_remainder'] or 0
        return total.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

    def get_free_usdt(self):
        total = self.get_sum_usdt() - self.get_sum_obligation()
        return total.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

    def get_bills_amount_dif(self):
        try:
            import requests
            url = 'https://api.rapira.net/open/market/rates'
            r = requests.get(url, timeout=2)
            r.raise_for_status()
            data = r.json()
            course_percent = PercentCourse.objects.first().value
            course_main = Decimal(data['data'][0]['close'])
            course_5 = course_main * (Decimal(1 + (course_percent / 100)))
            bills = Bill.objects.all()
            total_r = 0
            total_s = 0
            for bill in bills:
                if bill.currency.short_name in ['USD', 'USDT']:
                    total_r += bill.remainder * course_5
                    total_s += bill.rate * bill.remainder
                else:
                    total_s += bill.remainder
                    total_r += bill.remainder
            total = total_r - total_s
            return total.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        except:
            return 0


    def get_all_currencys(self):
        currencys = Currency.objects.all()
        return  currencys

    def get_courses(self, status):
        direction = ''
        variation = ''
        try:
            import requests
            url = 'https://api.rapira.net/open/market/rates'
            r = requests.get(url, timeout=2)
            r.raise_for_status()
            data = r.json()
            course_percent = PercentCourse.objects.first().value
            course_main = Decimal(data['data'][0]['close'])

            try:
                course_hist = CourseHistory.objects.get(currency__short_name='USDT')
                old_course = course_hist.new_value
                new_course = course_main
                if new_course > old_course:
                    direction = 'up'
                    variation = (new_course - old_course) / old_course * 100
                else:
                    direction = 'down'
                    variation = (new_course - old_course) / old_course * 100
                course_hist.old_value = old_course
                course_hist.new_value = new_course
                course_hist.save()
            except:
                currency = Currency.objects.get(short_name='USDT')
                course_hist = CourseHistory(currency=currency)
                course_hist.old_value = 0
                course_hist.new_value = course_main
                course_hist.save()
                direction = 'up'
                variation = 0
            course_5 = course_main * (Decimal(1 + (course_percent / 100)))
            course_tadg = course_5 + 3
            if status:
                return [{'name': 'Курс', 'value': course_5, 'direction': direction, 'variation': variation},
                        {'name': 'КурсТ', 'value': course_tadg, 'direction': direction, 'variation': variation}]
            else:
                return [
                    {'name': 'КурсA', 'value': course_5, 'direction': direction, 'variation': variation},
                    {'name': 'Курс', 'value': course_main, 'direction': direction, 'variation': variation},
                    {'name': 'КурсТ', 'value': course_tadg, 'direction': direction, 'variation': variation}
                ]
        except Exception as e:
            return [{'name': 'Курс', 'value': 'Недоступно', 'direction': direction,'variation': variation}]


    def get_dutys(self):
        list_dutys = []
        contractors = Contractor.objects.filter(status=True)#all()
        for contractor in contractors:
            contractor_data = {'name': contractor.name, 'initial_duty': contractor.duty}
            last_dutys = ContractorHistory.objects.filter(contractor=contractor)
            if last_dutys.count() != 0:
                if last_dutys.last().duty:
                    contractor_data['duty'] = float(last_dutys.last().duty)
                else:
                    contractor_data['duty'] = 0
                if last_dutys.last().duty_usdt:
                    contractor_data['duty_usdt'] = float(last_dutys.last().duty_usdt)
                else:
                    contractor_data['duty_usdt'] = 0
            else:
                contractor_data['duty'] = contractor.duty
                contractor_data['duty_usdt'] = contractor.duty_usdt
            list_dutys.append(contractor_data)

        return list_dutys

    def get_all_dutys(self):
        latest_hist = ContractorHistory.objects.filter(
            contractor=OuterRef('pk')
        ).order_by('-date_create')
        contractors_with_latest = Contractor.objects.annotate(
            latest_duty=Subquery(latest_hist.values('duty')[:1])
        )
        total = contractors_with_latest.aggregate(
            total_latest_duty=Sum('latest_duty')
        )['total_latest_duty'] or 0
        return Decimal(total).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)


    def changelist_view(self, request, extra_context=None):
        my_context = {
            'bills': self.get_bills(),
            'bills_amount': self.get_bills_amount(),
            'bills_amount_admin': self.get_bills_amount_admin(),
            'bills_amount_dif': self.get_bills_amount_dif(),
            'sum_obligation': self.get_sum_obligation(),
            'sum_usdt': self.get_sum_usdt(),
            'free_usdt': self.get_free_usdt(),
            'dutys_all': self.get_all_dutys(),
            'dutys': self.get_dutys(),
        }
        if request.user.is_superuser and request.user.groups.filter(name='restricted').exists():
            if request.user.id != 3:
                my_context['blocked'] = True
                my_context['is_admin'] = False
                my_context['courses'] = self.get_courses(True)
            else:
                my_context['blocked'] = False
                my_context['is_admin'] = False
                my_context['courses'] = self.get_courses(True)
        elif request.user.is_superuser:
            if request.user.id == 1:
                my_context['blocked'] = False
                my_context['is_admin'] = True
                my_context['courses'] = self.get_courses(False)
            else:
                my_context['blocked'] = False
                my_context['is_admin'] = False
                my_context['courses'] = self.get_courses(False)
        queryset = self.get_queryset(request)
        data = [
            {
                'id': obj.id,
                'custom': obj.my_color_data,
            } for obj in queryset
        ]
        in_exes = [
            
            {
                'id': obj.id,
                'income_before': float(obj.hb_income_before or 0),
                'income_after': float(obj.hb_income_after or 0),
                'expense_before': float(obj.hb_expense_before or 0),
                'expense_after': float(obj.hb_expense_after or 0),
                'national_currency_value': float(obj.national_currency_value or 0)
            } for obj in queryset
        ]
        data = json.dumps(data)
        my_context['color_data'] = data
        my_context['in_exes'] = in_exes
        return super(DealAdmin, self).changelist_view(request,
                                                       extra_context=my_context)

    def add_view(self, request, form_url='', extra_context=None):
        my_context = {
            'bills': self.get_bills(),
        }
        if request.user.is_superuser and request.user.groups.filter(name='restricted').exists():
            my_context['blocked'] = True
        else:
            my_context['blocked'] = False
        return super(DealAdmin, self).add_view(request, form_url, extra_context=my_context)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        my_context = {
            'bills': self.get_bills(),
        }
        if request.user.is_superuser and request.user.groups.filter(name='restricted').exists():
            my_context['blocked'] = True
        else:
            my_context['blocked'] = False
        deal = self.get_object(request, object_id)
        in_exes = IncomeExpense.objects.filter(deal=deal)
        in_exes_list = []
        for in_ex in in_exes:
            historys = HistoryBill.objects.filter(in_ex=in_ex)
            if historys.count() != 0:
                history_obj = historys.last()
                in_exes_list.append({
                    'id': in_ex.id,
                    'income_before': '{0} '.format(str(history_obj.income_before.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))),
                    'income_after': '{0} '.format(str(history_obj.income_after.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))),
                    'expense_before': '{0} '.format(str(history_obj.expense_before.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))),
                    'expense_after': '{0} '.format(str(history_obj.expense_after.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))),
                })
            else:
                in_exes_list.append({
                    'id': None,
                    'income_before': None,
                    'income_after': None,
                    'expense_before': None,
                    'expense_after': None,
                })
        my_context['in_exes'] = in_exes_list
        return super(DealAdmin, self).change_view(request, object_id, form_url, extra_context=my_context)

    def formatted_date_create(self, obj):
        local_dt = timezone.localtime(obj.date_create)
        date_st = local_dt.strftime('%d.%m.%y')
        date_full = local_dt.strftime('%d.%m.%y %H:%M:%S')
        return format_html(
            '<span class="js-toggle js-click-toggle" '
            'data-original="{}" data-alt="{}">{}</span>',
            date_full,
            'Больше?',
            date_st
        )


    def income_account_1(self, obj):
        in_exes =  IncomeExpense.objects.filter(deal=obj, income_account__isnull=False)
        if in_exes.count() != 0:
            in_ex = in_exes.first()
            if in_ex.income_account:
                return in_ex.income_account
        return '-'

    def income_amount_1(self, obj):
        in_exes =  IncomeExpense.objects.filter(deal=obj, income_amount__isnull=False)
        if in_exes.count() != 0:
            in_ex = in_exes.first()
            amount = in_ex.income_amount
            if amount and amount != 0:
                amount_in = '{0} '.format(str(amount.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)))
                return amount_in
        return '-'

    def expense_account_1(self, obj):
        in_exes = IncomeExpense.objects.filter(deal=obj, expense_account__isnull=False)
        if in_exes.count() != 0:
            in_ex = in_exes.first()
            if in_ex.expense_account:
                return in_ex.expense_account
        return '-'

    def expense_amount_1(self, obj):
        in_exes = IncomeExpense.objects.filter(deal=obj, expense_amount__isnull=False)
        if in_exes.count() != 0:
            in_ex = in_exes.first()
            amount = in_ex.expense_amount
            if amount and amount != 0:
                amount_in = '{0} '.format(str(amount.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)))
                return amount_in
        return '-'

    def created_by(self, obj):
        ct = ContentType.objects.get_for_model(obj)
        entry = (
            LogEntry.objects
            .filter(content_type=ct,
                    object_id=str(obj.pk),
                    action_flag=ADDITION)
            .order_by('action_time')
            .first()
        )
        return entry.user.username if entry else ''

    def national_currency_custom(self, obj):
        if obj.national_currency:
            return obj.national_currency.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        return ''


    national_currency_custom.allow_tags = True
    national_currency_custom.short_description = 'Нац.валюта'
    created_by.short_description = 'Кем создано'
    formatted_date_create.short_description = 'Дата'
    formatted_date_create.admin_order_field = 'date_create'
    income_account_sum.allow_tags = True
    income_account_sum.short_description = "Счет прихода/сумма"
    expense_account_sum.allow_tags = True
    expense_account_sum.short_description = "Счет расхода/сумма"
    income_account_1.allow_tags = True
    income_account_1.short_description = "Счет прихода"
    expense_account_1.allow_tags = True
    expense_account_1.short_description = "Счет расхода"
    income_amount_1.allow_tags = True
    income_amount_1.short_description = "Сумма прихода"
    expense_amount_1.allow_tags = True
    expense_amount_1.short_description = "Сумма расхода"


